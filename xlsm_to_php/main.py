import openpyxl
import os


def associa_colonne():
    current_dir = os.getcwd()
    file_path = os.path.join(current_dir, "mci-expression-list.xlsm")

    workbook = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True, data_only=True)
    worksheet = workbook.worksheets[0]

    data = {}
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        key = row[0]
        for col_index, cell_value in enumerate(row[1:], start=2):
            if key not in data:
                data[key] = {}
            data[key][f"col{col_index}"] = cell_value

    output_dir = current_dir

    for col_index in range(2, worksheet.max_column + 1):
        json_data = {}
        for key, values in data.items():
            if f"col{col_index}" in values:
                json_data[key] = values[f"col{col_index}"]
            else:
                json_data[key] = None

        if all(value is None for value in json_data.values()):
            break

        php_file_path = os.path.join(output_dir, "mci-expression-list-{}.php".format(worksheet.cell(row=2, column=col_index).value))
        generate_php_file(json_data, php_file_path)

        print("File generated: {}".format(php_file_path) + "\n")


def generate_php_file(data, output_file):
    with open(output_file, 'w', encoding='utf-8-sig') as php_file:
        php_file.write("<?php\n")
        for key, value in data.items():
            escaped_value = value.replace('\\', '\\\\').replace('"', '\\"') if value is not None else ""
            php_file.write(f'${key} = "{escaped_value}";\n')
        php_file.write("?>")


print("This script will output a PHP file.\n" +
"The PHP file will be a list that declares the key (the first column of the “xlsm” file) as a variable,"
" and assigns it the value (the contents of the cells of the other columns).\n\n")
print("WARNING! The xlsm file must necessarily be in the same folder as the xlsm_to_php.exe executable, and the file xlsm MUST BE named as -> mci-expression-list.xlsm\n\n")


input("Press Enter to start generating the files...")
associa_colonne()
print("File generation completed.")
input("Press Enter to exit.")
