import os
import shutil
from openpyxl import load_workbook

def copy_files_from_xlsm(file_xlsm):
    # Leggi il file XLSM
    workbook = load_workbook(file_xlsm)
    sheet = workbook.active

    # Ottieni il percorso della cartella contenente il file XLSM
    xlsm_path = os.path.dirname(os.path.abspath(file_xlsm))

    # Percorso della cartella di destinazione
    destination_folder = os.path.join(xlsm_path, 'file_di_esportazione')

    # Crea la cartella di destinazione se non esiste
    os.makedirs(destination_folder, exist_ok=True)

    # Leggi i dati dal file XLSM e copia i file e le cartelle
    for row in sheet.iter_rows(values_only=True):
        file_name = row[0]
        action = row[1]

        if file_name:
            # Percorso completo del file di origine
            source_file = os.path.join(xlsm_path, file_name)

            if os.path.exists(source_file):
                if action == 'CREA CARTELLA':
                    # Crea una cartella vuota nella cartella di destinazione
                    new_folder = os.path.join(destination_folder, file_name)
                    os.makedirs(new_folder, exist_ok=True)
                elif action == 'CARTELLA CON IL SUO CONTENUTO':
                    # Copia la cartella e tutto il suo contenuto nella cartella di destinazione
                    shutil.copytree(source_file, os.path.join(destination_folder, file_name))
                else:
                    if '/' in file_name:
                        # Se il nome file contiene il carattere '/', significa che è specificato un percorso
                        destination_folder_name, file_name_only = file_name.rsplit('/', 1)
                        destination_folder_path = os.path.join(destination_folder, destination_folder_name)
                        os.makedirs(destination_folder_path, exist_ok=True)
                        shutil.copy2(source_file, os.path.join(destination_folder_path, file_name_only))
                    else:
                        # Copia solo il file nella cartella di destinazione
                        shutil.copy2(source_file, os.path.join(destination_folder, file_name))

xlsm_file = 'file_da_esportare.xlsm'
copy_files_from_xlsm(xlsm_file)
