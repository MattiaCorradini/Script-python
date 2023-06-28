import openpyxl
import json


def read_xlsm_file(file_path, start_column):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    dati = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue

        mec = row[0]
        errortype = row[1]
        confirmability = row[2]
        detail_id = row[3]

        if start_column < len(row):
            text = row[start_column]
        else:
            break

        if mec not in dati:
            dati[mec] = {
                "errortype": errortype,
                "confirmability": confirmability,
                "detail-ID": detail_id,
                "text": text
            }

    return dati


def write_json_file(dati, file_path):
    with open(file_path, 'w') as json_file:
        json.dump(dati, json_file, indent=4)


xlsm_file_path = 'mci-errors.xlsm'
start_column_index = 4

workbook = openpyxl.load_workbook(xlsm_file_path)
sheet = workbook.active

print("This script will output JSON files.")
print("The JSON files will be an object that contains some key-values associations, one file for each language.\n")
print("Each key-value pair in the JSON object represents an error code and its associated information. Here's the breakdown of each field:")
print("errortype: Indicates the type of error (e.g., error or warning)")
print("confirmability: Specifies whether the error is confirmable or not confirmable.")
print("detail-ID: An identifier for additional details of the error. It can be null if there are no specific details provided.")
print("text: A description of the error in the language indicated by the name of the file.\n\n")
print("WARNING! The xlsm file must necessarily be in the same folder as the emcylist-CONVERTER.exe executable, and the file xlsm MUST BE named as -> emcylist.xlsm\n\n")

input("Press Enter to start generating the files...")
for column_index in range(start_column_index, sheet.max_column + 1):
    column_letter = openpyxl.utils.get_column_letter(column_index+1)
    first_row_value = sheet[column_letter + '1'].value

    if not first_row_value:
        continue

    json_file_name = f'mci-errors-{first_row_value.upper()}.json'
    data = read_xlsm_file(xlsm_file_path, column_index)
    write_json_file(data, json_file_name)
    print("File generated: {}".format(json_file_name) + "\n")
print("File generation completed.")
input("Press Enter to exit.")