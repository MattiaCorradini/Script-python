import time
import openpyxl
import json
import os
import sys


def associate_columns(path_file):
    workbook = openpyxl.load_workbook(path_file, read_only=True, keep_vba=True, data_only=True)
    worksheet = workbook.worksheets[0]

    data = {}
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        key = row[0]
        for col_index, cell_value in enumerate(row[1:], start=2):
            if key not in data:
                data[key] = {}
            data[key][f"col{col_index}"] = cell_value

    output_dir = os.getcwd()

    php_dir = os.path.join(output_dir, "PHP")
    json_dir = os.path.join(output_dir, "JSON")

    os.makedirs(php_dir, exist_ok=True)
    os.makedirs(json_dir, exist_ok=True)

    for col_index in range(2, worksheet.max_column + 1):
        if not any(data[key].get(f"col{col_index}") for key in data):
            break

        json_data = {}
        for key, values in data.items():
            value = values.get(f"col{col_index}")
            if value is None:
                value = ""
            json_data[key] = value

        json_data_str = json.dumps(json_data, indent=4)

        json_file_path = os.path.join(json_dir, "{}_{}.json".format(os.path.splitext(os.path.basename(path_file))[0], worksheet.cell(row=2, column=col_index).value))
        with open(json_file_path, "w") as json_file:
            json_file.write(json_data_str)

        php_file_path = os.path.join(php_dir, "{}_{}.php".format(os.path.splitext(os.path.basename(path_file))[0], worksheet.cell(row=2, column=col_index).value))
        generate_php_file(json_file_path, php_file_path)

        print("Generated files: {} and {}".format(json_file_path, php_file_path) + "\n")


def generate_php_file(input_file, output_file):
    with open(input_file, 'r', encoding='utf-8-sig') as json_file, open(output_file, 'w', encoding='utf-8-sig') as php_file:
        data = json.load(json_file)
        php_file.write("<?php\n")
        for key, value in data.items():
            escaped_value = value.replace('\\', '\\\\').replace('"', '\\"') if value is not None else ""
            php_file.write(f'${key} = "{escaped_value}";\n')
        php_file.write("?>")


default_file_path = "MCI_expression_list.xlsm"
file_path = None

print("APPLICATION THAT CONVERTS AN XLSM FILE INTO JSON AND PHP"", " "\nAND PUT THEM INTO THEIR REPOSITORY IN THE SAME FOLDER AS THE EXE AND XLSM FILES\n")

if os.path.isfile(default_file_path):
    print("The file '{}' was found.".format(default_file_path) + "\n")
    choice = input("Do you want to use this file for conversion? (YES/NO): ")
    if choice.lower() == 'yes':
        file_path = default_file_path


if file_path is None or not os.path.isfile(file_path):
    print("\nDrag and drop the Excel file here and press Enter to start the conversion: ")
    file_path = input().strip()

if not os.path.isfile(file_path):
    print("\nThe file '{}' does not exist. Exiting...".format(file_path))
    time.sleep(3)
    sys.exit(1)

if not file_path.lower().endswith(".xlsm"):
    print("\nThe selected file does not have the .xlsm extension. Invalid file, exiting...")
    time.sleep(3)
    sys.exit(1)


print("\nUsing file: {}".format(file_path) + "\n")
input("Press Enter to start generating the files..." + "\n")
associate_columns(file_path)
print("File generation completed.")

input("Press Enter to exit.")
