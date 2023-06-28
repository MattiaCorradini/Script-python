import openpyxl
import json


def associa_colonne(nome_file):
    workbook = openpyxl.load_workbook(nome_file, read_only=True, keep_vba=True, data_only=True)

    worksheet = workbook.worksheets[0]

    data = {}  # Dizionario per memorizzare le associazioni tra la prima colonna e le altre colonne
    for row in worksheet.iter_rows(values_only=True):
        key = row[0]  # Contenuto della prima colonna (chiave)
        for col_index, cell_value in enumerate(row[1:], start=2):
            if key not in data:
                data[key] = {}
            data[key][f"col{col_index}"] = cell_value

    # Salva i file JSON per le associazioni tra la prima colonna e le colonne successive
    for col_index in range(2, worksheet.max_column + 1):
        json_data = {}
        for key, values in data.items():
            if f"col{col_index}" in values:
                json_data[key] = values[f"col{col_index}"]
            else:
                json_data[key] = None

        json_data_str = json.dumps(json_data, indent=4)
        json_file_path = nome_file.replace("es.xlsm",
                                           f"expression-list-{worksheet.cell(row=2, column=col_index).value}.json")
        with open(json_file_path, "w") as json_file:
            json_file.write(json_data_str)


file_name = "ESEMPIO EXCEL.xlsm"
associa_colonne(file_name)
